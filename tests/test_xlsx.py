from datetime import date
from pathlib import Path
import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import sessionmaker
from app.db import Base, make_engine
from app import models as m
from app.settings import Settings
from toolkit import ROOT
from toolkit.xlsx.export import build_workbook
from toolkit.xlsx.import_tracker import import_tracker
from toolkit.xlsx.check import check

TRACKER = ROOT / "Petition Captain Master Tracker.xlsx"


@pytest.fixture
def db():
    eng = make_engine("sqlite://"); Base.metadata.create_all(eng)
    with sessionmaker(bind=eng, expire_on_commit=False)() as s:
        yield s


def _sample(db):
    c = m.Circulator(name="Ada Lovelace", role="Circulator", registered_voter_verified=True, trained_on=date(2026, 9, 1))
    db.add(c)
    p = m.Pamphlet(number="P-001", status="Returned", issued_to=c, issued_on=date(2026, 9, 2), returned_on=date(2026, 9, 3))
    p.sheets = [m.Sheet(sheet_no=i, status="Notarized", collected=10, questionable=1, rejected=0, returned_on=date(2026, 9, 3),
                        notarized_on=date(2026, 9, 3), notary_name="N. Public", notary_commission="12345", notary_expiration=date(2028, 1, 1),
                        defect_codes="E7" if i == 5 else None) for i in range(1, 6)]
    db.add(p); db.add(m.Pamphlet(number="P-002", sheets=[m.Sheet(sheet_no=i) for i in range(1, 6)]))
    db.add(m.Issue(number="I-001", status="Open", issue_type="E7", pamphlet=p, opened_on=date(2026, 9, 4)))
    db.add(m.QATask(task="Do the thing", status="Done", sort_order=10))
    loc = m.Location(slug="stipe", name="J.I. Stipe Center", city="McAlester"); db.add(loc)
    db.add(m.Event(location=loc, date=date(2026, 9, 6), lead=c, volunteers_needed=3))
    db.add(m.Contact(role="Petition Captain", name="Ada Lovelace", phone="918-555-0100"))
    db.add(m.RecordsLog(item="True copy filed", office="County Election Board", receipt_obtained=True))
    s = Settings(db); s.set("registered_voters", 27590); s.set("adoption_date", date(2026, 10, 5))
    db.commit()


def test_export_from_db(db, tmp_path):
    _sample(db)
    wb = build_workbook(db)
    out = tmp_path / "pm.xlsx"; wb.save(out)
    wb = load_workbook(out)
    assert wb.sheetnames[:3] == ["Dashboard", "README", "Pamphlet Log"]
    for name in ["Signature Sheets", "Daily Counts", "Volunteers", "Filing QA", "EventsShifts", "Issues", "Deadlines", "Notary Log",
                 "Records Log", "Contacts", "Precincts", "Defect Codes", "RulesSources", "Dropdowns"]:
        assert name in wb.sheetnames
    pl, ss, d = wb["Pamphlet Log"], wb["Signature Sheets"], wb["Dashboard"]
    assert [c.value for c in pl[1]][:14] == ["Pamphlet #", "Status", "Printed Date", "Issued To", "Issued Date", "Returned Date", "Sheet Count",
                                              "Expected Capacity", "Collected Count", "Notarized Sheets", "Audited OK Sheets", "Rejected Sheets", "Filed?", "Notes"]
    assert pl["A2"].value == "P-001" and pl["B2"].value == "Returned" and pl["D2"].value == "Ada Lovelace"
    assert pl["I2"].value.startswith("=SUMIFS('Signature Sheets'") and pl["M2"].value == '=IF(B2="Filed","Yes","No")'
    assert pl.max_row == 3                                   # DB has 2 pamphlets → no template padding
    assert ss["C2"].value == "P-001-S1" and ss["L2"].value == "=MAX(I2-J2-K2,0)" and ss["I2"].value == 10
    assert ss["Q6"].value == "E7" and ss["N2"].value == "N. Public"
    assert d["B7"].value == '=IF(B13="","",ROUNDUP(B13*B14,0))' and d["B13"].value == 27590 and d["B14"].value == 0.10
    assert d["B17"].value.date() == date(2026, 10, 5) and d["B18"].value == '=IF(B17="","",B17+30)'
    assert d["E11"].value == 50 and d["H11"].value == 45      # DB cross-check from signature_stats
    assert "AdoptionDate" in wb.defined_names and wb.defined_names["AdoptionDate"].attr_text == "Dashboard!$B$17"
    dc = wb["Daily Counts"]
    assert dc["A2"].value == '=IF(AdoptionDate="","",AdoptionDate+0)' and "SUMIFS('Signature Sheets'!$I:$I" in dc["B2"].value
    assert wb["Volunteers"]["A2"].value == "Ada Lovelace" and wb["Volunteers"]["E2"].value == "Yes"
    assert wb["Issues"]["A2"].value == "I-001" and wb["Issues"]["E2"].value == "E7" and wb["Issues"]["A301"].value == "I-300"
    assert wb["Issues"]["F3"].value is None                    # unused issue rows carry no status → no phantom open issues
    assert wb["Precincts"].max_row >= 39 and wb["Precincts"]["A2"].value == 1
    assert wb["Defect Codes"]["A2"].value == "E1" and "notary" in wb["Defect Codes"]["B8"].value
    assert wb["RulesSources"]["B2"].value == "62 O.S. § 868" and "oklegislature" in wb["RulesSources"]["E2"].value
    assert wb["Deadlines"]["D7"].value == '=IF(FilingDeadline="","",FilingDeadline)'
    assert wb["Notary Log"]["F2"].value == "N. Public"
    assert wb["Records Log"]["B2"].value == "True copy filed"
    assert wb["Contacts"]["B2"].value == "Ada Lovelace"
    assert "SYSTEM OF RECORD" in wb["README"]["A4"].value
    for ws in wb.worksheets:
        assert ws.page_setup.paperSize == 5 and ws.page_setup.orientation == "landscape"
    assert check(str(out)) == []


def test_export_blank_template(db, tmp_path):
    wb = build_workbook(db)
    out = tmp_path / "blank.xlsx"; wb.save(out)
    wb = load_workbook(out)
    assert wb["Pamphlet Log"].max_row == 201 and wb["Pamphlet Log"]["A201"].value == "P-200"
    assert wb["Signature Sheets"].max_row == 1001 and wb["Signature Sheets"]["C1001"].value == "P-200-S5"
    assert wb["Dashboard"]["B13"].value == 27727 and wb["Dashboard"]["B17"].value is None   # voters from config; adoption still TBD
    assert wb["Filing QA"]["A2"].value.startswith("Get the exact adopted resolution")
    assert check(str(out)) == []


@pytest.mark.skipif(not TRACKER.exists(), reason="legacy tracker not present")
def test_import_legacy_tracker(db, tmp_path):
    c1 = import_tracker(TRACKER, db)
    assert c1["pamphlets"] == 200 and c1["sheets"] == 1000 and c1["issues"] == 0 and c1["qa_tasks"] == 13
    assert any("NOT imported" in w for w in c1["warnings"])
    assert db.get(m.Setting, "registered_voters") is None          # import never writes the voter count
    s = Settings(db)
    assert s.registered_voters == 27727 and s.legal_minimum == 2773  # falls through to config/petition.yaml
    assert s.print_run == 200 and s.sheets_per_pamphlet == 5 and s.rows_per_sheet == 10 and s.est_valid_rate == 0.85
    assert s.float("overcollect_fraction") == 0.3
    assert db.query(m.Pamphlet).count() == 200 and db.query(m.Sheet).count() == 1000
    assert db.get(m.Pamphlet, 1).number == "P-001" and db.get(m.Pamphlet, 1).status == "Ready to Print"
    c2 = import_tracker(TRACKER, db)                          # idempotent
    assert c2["pamphlets"] == 0 and c2["sheets"] == 0 and c2["issues"] == 0
    assert db.query(m.Pamphlet).count() == 200 and db.query(m.Issue).count() == 0
    wb = build_workbook(db); out = tmp_path / "after-import.xlsx"; wb.save(out)
    assert check(str(out)) == []
    wb = load_workbook(out)
    assert wb["Pamphlet Log"]["A201"].value == "P-200" and wb["Signature Sheets"].max_row == 1001


def test_check_catches_problems(tmp_path):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "Dashboard"; ws["A7"] = "Legal Minimum Needed"; ws["B7"] = 2759; ws["A8"] = "deadline July 22"
    ws["C1"] = "=Nope!A1"
    out = tmp_path / "bad.xlsx"; wb.save(out)
    errs = check(str(out))
    assert any("missing sheet" in e for e in errs) and any("2759" in e for e in errs) and any("July 22" in e for e in errs)
    assert any("unknown sheet 'Nope'" in e for e in errs) and any("formula" in e for e in errs)
