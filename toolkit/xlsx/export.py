"""Export the Petition Master workbook from the database.

The app database is the system of record. This workbook is a faithful, offline-usable
export that keeps the sheet names, column order, dropdowns, conditional formatting and
live Excel formulas of the captain's original "Petition Captain Master Tracker.xlsx",
and adds statutory reference sheets. It never contains signer names, addresses or
birth dates — the model stores none.

    python -m toolkit.xlsx.export --out output/xlsx/petition-master.xlsx [--database-url URL]
"""
from __future__ import annotations
import argparse, csv, json
from datetime import date, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import select
from sqlalchemy.orm import Session
from toolkit import ROOT, config as cfg, statutes
from app import models as m
from app.settings import Settings
from app.stats import signature_stats

HEADER_FILL = PatternFill("solid", fgColor="173A73")
HEADER_FONT = Font(bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")      # yellow = captain enters a value here
CALC_FILL = PatternFill("solid", fgColor="EEF2F7")       # grey-blue = formula, do not type over
THIN = Side(style="thin", color="C9CED6")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DATE_FMT = "yyyy-mm-dd"
LEGAL = 5  # openpyxl PAPERSIZE_LEGAL

# Column order of the original tracker, extended (new columns only ever go on the END).
COLS = {
    "Pamphlet Log": ["Pamphlet #", "Status", "Printed Date", "Issued To", "Issued Date", "Returned Date", "Sheet Count",
                     "Expected Capacity", "Collected Count", "Notarized Sheets", "Audited OK Sheets", "Rejected Sheets",
                     "Filed?", "Notes", "Print Batch", "Version Hash"],
    "Signature Sheets": ["Pamphlet #", "Sheet #", "Sheet ID", "Circulator", "Issued Date", "Returned Date", "Notarized Date",
                         "Status", "Collected Signatures", "Questionable", "Rejected", "Valid Estimate", "Notes",
                         "Notary Name", "Notary Commission #", "Notary Expiration", "Defect Codes"],
    "Daily Counts": ["Date", "Collected Signatures", "Questionable", "Rejected", "Valid Estimate", "Pamphlets Returned", "Notes", "Event/Location"],
    "Volunteers": ["Name", "Role", "Phone", "Email", "County Registered Voter?", "Training Complete?", "Notary?",
                   "Assigned Pamphlets", "Availability", "Notes", "Verified On", "Verified By", "Trained On", "Compensated?", "Active?"],
    "Filing QA": ["Task", "Status", "Owner", "Notes"],
    "EventsShifts": ["Date", "Start", "End", "Location", "Event Lead", "Volunteers Needed", "Pamphlets Issued", "Expected Signatures", "Notes"],
    "Issues": ["Issue #", "Date", "Pamphlet #", "Sheet ID", "Issue Type", "Status", "Priority", "Resolution / Notes"],
    "Deadlines": ["Step", "Statute", "Rule", "Date", "How the date is set", "Status / Notes"],
    "Notary Log": ["Date", "Pamphlet #", "Sheet #", "Circulator", "Sig. Count", "Notary Name", "Commission #", "Expiration", "Notary Initials", "Notes"],
    "Records Log": ["Date / Time", "Item", "Office", "Person", "Documents", "Receipt Obtained?", "Notes"],
    "Contacts": ["Role", "Name", "Phone", "Email", "Address", "Hours", "Public?"],
    "Precincts": ["Precinct", "Polling Place", "Address", "City", "2020 Population", "Voting-Age Pop (18+)", "Commissioner District", "State House", "State Senate"],
    "Defect Codes": ["Code", "34 O.S. § 6.1(A) — signatures not counted", "Scope", "How to prevent it"],
    "RulesSources": ["Topic", "Cite", "Title", "What it means for us", "Source URL", "Retrieved"],
}
DEFAULT_QA_TASKS = [
    "Get the exact adopted resolution and all attachments after the Board adopts it (tabled — date TBD)",
    "Insert exact resolution number, title, and adopted text into the pamphlet",
    "Obtain the written county registered voter count from the Election Board and set the 10% threshold",
    "Finalize petition text, gist, and ballot title (≤150 words — 62 O.S. § 868(D))",
    "Legal review of wording and packet layout by an Oklahoma election-law attorney",
    "File true copy with Secretary of County Election Board before circulation (62 O.S. § 868(B)(1))",
    "File the ballot title with the true copy so the DA's 3-day review runs early (62 O.S. § 868(D)(2))",
    "Get file-stamped copy with date/time/person accepting filing",
    "Print only the final filed version; record the version hash on every print batch",
    "Number all pamphlets P-001 through P-{run:03d}",
    "Verify every circulator is a registered Oklahoma voter before issuing a pamphlet (34 O.S. § 6)",
    "Train all circulators before issuing packets",
    "Track each pamphlet issued and returned",
    "Audit all sheets for missing fields, notarization and E1–E8 defects",
    "File signed petitions before the 30-day deadline (62 O.S. § 868(B)(3))",
]
DEFAULT_RECORDS = [
    "Resolution adopted (number, title, exact vote, date, amendments)",
    "Certified / file-stamped copy of adopted resolution requested",
    "Registered voter count obtained (total, date, source, 10% minimum)",
    "True copy of petition filed before circulation (date/time, office, receiving person, file-stamped copy retained)",
    "Ballot title filed",
    "Petition pamphlets printed (date, version hash, copies, who controls distribution)",
    "Signed petitions submitted (date/time, office, receiving person, number of pamphlets, receipt)",
]
DEFECT_SCOPE = {"E1": "WHOLE SHEET", "E2": "signature", "E3": "WHOLE SHEET", "E4": "signature(s) on that line",
                "E5": "signature", "E6": "signature", "E7": "WHOLE SHEET", "E8": "signature"}
DEFECT_PREVENT = {
    "E1": "Circulator signs the affidavit only after collecting, in front of the notary. Never a blank affidavit.",
    "E2": "Ask: registered voter living in Pittsburg County? Point non-residents to the Voter Portal instead of signing.",
    "E3": "Never detach, reorder or replace pages. A pamphlet stays a pamphlet.",
    "E4": "One person per numbered line. Cross out and re-do on a fresh line if two people share one.",
    "E5": "Signatures go only on printed, numbered lines — never in margins or between rows.",
    "E6": "Signer uses their own legal name; nobody signs for a spouse/parent/friend; no repeat signers.",
    "E7": "Notary checklist before leaving the table: signature, seal, commission number, expiration date, correct venue.",
    "E8": "The five fields (legal first, legal last, ZIP, house #, birth MM/DD) must match the voter registration; check legibility before the signer leaves.",
}
STATUTE_MEANING = {
    "62-868": "County referendum on Local Development Act incentives: true copy filed with the Secretary of the County Election Board BEFORE circulation; 10% of registered county voters; signed copies within 30 days of adoption; 10-day protest window; ballot title ≤150 words with yes/no language, DA review in 3 days, 10-day appeal; vote at the next general county election.",
    "34-1": "The petition form. Signer attestation ('each for himself says…'); the five data points (legal first name, legal last name, ZIP, house number, birth month/day); 4 of 5 must match the voter file; 30 days for county measures.",
    "34-3": "Every signature sheet is attached to a full copy of the petition (a pamphlet); 'Warning' + felony sentence on the outer page in ≥10-pt type; neutral gist on the top margin of each sheet; Open Records notice under it.",
    "34-6": "Circulators must be registered Oklahoma voters and swear the affidavit printed on the back of each signature sheet before an Oklahoma notary (signature, title, address, commission number, expiration, seal).",
    "34-6.1": "The eight grounds on which signatures are thrown out (defect codes E1–E8). E1, E3 and E7 lose the entire sheet.",
    "34-23": "Only qualified electors may sign; signing another name, signing twice, or signing while not a legal voter is a Class D3 felony.",
    "34-24": "Substantial compliance suffices and clerical errors are disregarded — a safety net, never a plan.",
    "34-2": "Initiative petition form (90-day window) — relevant only if proponents try to initiate the abatement by petition.",
    "34-9": "State ballot-title procedure: filed on a separate sheet, not printed on the petition; neutral; yes/no clarity.",
}


def _style_header(ws, ncols: int, row: int = 1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


def _widths(ws, widths: dict[str, float]):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _print_setup(ws, title_rows="1:1"):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = LEGAL
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.gridLines = True
    ws.page_margins.left = ws.page_margins.right = 0.5
    if title_rows:
        ws.print_title_rows = title_rows
    ws.oddFooter.center.text = "&A — Page &P of &N — exported &D"


def _table(ws, name: str, ref: str):
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(t)


def _cf(ws, rng: str, formula: str, rgb: str):
    ws.conditional_formatting.add(rng, FormulaRule(formula=[formula], fill=PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")))


def _dv(ws, formula1: str, sqref: str, allow_blank=True):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=allow_blank)
    ws.add_data_validation(dv)
    dv.add(sqref)
    return dv


def _yes(b) -> str:
    return "Yes" if b else "No"


def _sheet(wb: Workbook, title: str, headers: list[str], widths: dict[str, float] | None = None, freeze="A2"):
    ws = wb.create_sheet(title)
    ws.append(headers)
    _style_header(ws, len(headers))
    if widths: _widths(ws, widths)
    if freeze: ws.freeze_panes = freeze
    _print_setup(ws)
    return ws


def _q(name: str) -> str:
    return f"'{name}'" if " " in name else name


# --------------------------------------------------------------------------- data sources
def _precincts() -> list[dict]:
    places = {r["precinct"]: r for r in csv.DictReader(open(ROOT / "data" / "polling_places.csv", encoding="utf-8"))}
    g = json.load(open(ROOT / "data" / "precincts" / "pittsburg_pct2020.geojson", encoding="utf-8"))
    rows = []
    for f in g["features"]:
        p = f["properties"]; n = str(p.get("Precinct") or p.get("precinct_num"))
        pl = places.get(n, {})
        rows.append({"precinct": int(n), "polling_place": pl.get("polling_place"), "address": pl.get("address"),
                     "city": pl.get("city"), "pop": int(p.get("P0010001") or 0), "vap": int(p.get("P0030001") or 0),
                     "comm": int(p["Comm"]) if p.get("Comm") else None, "house": p.get("St_house"), "senate": p.get("St_senate")})
    return sorted(rows, key=lambda r: r["precinct"])


# --------------------------------------------------------------------------- builder
def build_workbook(db: Session, settings: Settings | None = None, petition: cfg.Petition | None = None) -> Workbook:
    petition = petition or cfg.load()
    s = settings or Settings(db, petition)
    stats = signature_stats(db, s)
    now = datetime.now()
    run, spp, rps = s.print_run, s.sheets_per_pamphlet, s.rows_per_sheet

    pamphlets = db.scalars(select(m.Pamphlet).order_by(m.Pamphlet.number)).all()
    template_mode = not pamphlets
    wb = Workbook()
    wb.remove(wb.active)

    # ---- README ---------------------------------------------------------------------
    ws = wb.create_sheet("README")
    lines = [
        ("Petition Master workbook — Pittsburg County referendum", 14, True),
        (f"Exported {now:%Y-%m-%d %H:%M} from the petition.mcalester.net database" + (" (EMPTY DATABASE → blank template)" if template_mode else ""), 11, False),
        ("", 11, False),
        ("THE APP DATABASE (admin area) IS THE SYSTEM OF RECORD. This file is an export for offline use, printing and backup.", 11, True),
        ("Edits made here are NOT read back automatically. Enter the same facts in the admin area (or ask for an import).", 11, False),
        ("", 11, False),
        ("THIS FILE MUST NEVER HOLD SIGNER NAMES, ADDRESSES OR BIRTH DATES. It tracks pamphlets and sheets by count only.", 11, True),
        ("Signer information exists only on the paper pamphlets. Do not photograph signed sheets. (34 O.S. § 3(B): they are public records once filed — until then, protect them.)", 11, False),
        ("", 11, False),
        ("Colour key: yellow = value you enter; grey-blue = formula, do not type over; dropdown cells only accept the listed statuses (see Dropdowns sheet).", 11, False),
        ((f"Registered voters: {s.registered_voters:,} (source: {s.raw('registered_voters_source') or 'unrecorded'}, as of {s.raw('registered_voters_date') or 'unknown'}) → legal minimum {s.legal_minimum:,} (10%). Keep a written, dated copy in the Records Log."
          if s.registered_voters else "Placeholder: the registered-voter count is BLANK until the County Election Board provides a written figure."), 11, False),
        ("Placeholder: the Adoption Date is BLANK until the Board adopts the resolution (currently tabled). Nothing may be circulated before adoption AND filing of the true copy.", 11, False),
        ("Every date on the Deadlines and Daily Counts sheets is computed from the Adoption Date cell on the Dashboard — set it once, everything follows (62 O.S. § 868).", 11, False),
        ("Sheets: Dashboard · Pamphlet Log · Signature Sheets · Daily Counts · Volunteers · Filing QA · EventsShifts · Issues · Deadlines · Notary Log · Records Log · Contacts · Precincts · Defect Codes · RulesSources · Dropdowns", 11, False),
    ]
    for i, (txt, size, bold) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=txt); c.font = Font(size=size, bold=bold); c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 18 if size == 11 else 24
    ws.column_dimensions["A"].width = 140
    _print_setup(ws, title_rows=None)

    # ---- Dropdowns (built before the sheets that reference it) ----------------------
    dd = wb.create_sheet("Dropdowns")
    dd_cols = [("Pamphlet Status", m.PAMPHLET_STATUSES), ("Sheet Status", m.SHEET_STATUSES), ("Volunteer Role", m.VOLUNTEER_ROLES),
               ("Issue Status", m.ISSUE_STATUSES), ("Priority", m.ISSUE_PRIORITIES), ("Issue Type", m.ISSUE_TYPES),
               ("QA Status", m.QA_STATUSES), ("Location Status", m.LOCATION_STATUSES), ("Yes/No", ["Yes", "No"]), ("Defect Codes", m.DEFECT_CODES)]
    dd_ref: dict[str, str] = {}
    for ci, (name, items) in enumerate(dd_cols, 1):
        col = get_column_letter(ci)
        dd.cell(row=1, column=ci, value=name)
        for ri, it in enumerate(items, 2):
            dd.cell(row=ri, column=ci, value=it)
        dd_ref[name] = f"Dropdowns!${col}$2:${col}${len(items) + 1}"
        dd.column_dimensions[col].width = 22
    _style_header(dd, len(dd_cols))
    _print_setup(dd)

    # ---- Pamphlet Log ---------------------------------------------------------------
    pl = _sheet(wb, "Pamphlet Log", COLS["Pamphlet Log"],
                {"A": 12, "B": 16, "C": 13, "D": 30, "E": 13, "F": 13, "G": 11, "H": 12, "I": 12, "J": 12, "K": 13, "L": 12, "M": 8, "N": 32, "O": 14, "P": 20})
    pam_rows = [(p.number, p.status, p.printed_on, p.issued_to.name if p.issued_to else None, p.issued_on, p.returned_on,
                 len(p.sheets) or spp, p.notes, p.print_batch, p.version_hash) for p in pamphlets] if not template_mode else \
               [(f"P-{i:03d}", "Ready to Print", None, None, None, None, spp, None, None, None) for i in range(1, run + 1)]
    for r, (num, st, pd, who, iss, ret, cnt, notes, batch, vh) in enumerate(pam_rows, 2):
        pl.cell(r, 1, num); pl.cell(r, 2, st); pl.cell(r, 3, pd); pl.cell(r, 4, who); pl.cell(r, 5, iss); pl.cell(r, 6, ret)
        pl.cell(r, 7, cnt); pl.cell(r, 8, f"=G{r}*Dashboard!$B$5")
        pl.cell(r, 9, f"=SUMIFS('Signature Sheets'!$I:$I,'Signature Sheets'!$A:$A,A{r})")
        pl.cell(r, 10, f"=COUNTIFS('Signature Sheets'!$A:$A,A{r},'Signature Sheets'!$H:$H,\"Notarized\")+COUNTIFS('Signature Sheets'!$A:$A,A{r},'Signature Sheets'!$H:$H,\"Audited OK\")+COUNTIFS('Signature Sheets'!$A:$A,A{r},'Signature Sheets'!$H:$H,\"Filed\")")
        pl.cell(r, 11, f"=COUNTIFS('Signature Sheets'!$A:$A,A{r},'Signature Sheets'!$H:$H,\"Audited OK\")+COUNTIFS('Signature Sheets'!$A:$A,A{r},'Signature Sheets'!$H:$H,\"Filed\")")
        pl.cell(r, 12, f"=COUNTIFS('Signature Sheets'!$A:$A,A{r},'Signature Sheets'!$H:$H,\"Rejected\")")
        pl.cell(r, 13, f"=IF(B{r}=\"Filed\",\"Yes\",\"No\")"); pl.cell(r, 14, notes); pl.cell(r, 15, batch); pl.cell(r, 16, vh)
        for c in (3, 5, 6): pl.cell(r, c).number_format = DATE_FMT
        for c in (8, 9, 10, 11, 12, 13): pl.cell(r, c).fill = CALC_FILL
    pl_last = len(pam_rows) + 1
    _table(pl, "Table_2", f"A1:P{pl_last}")
    _dv(pl, dd_ref["Pamphlet Status"], f"B2:B{pl_last}")
    _dv(pl, "=Volunteers!$A$2:$A$201", f"D2:D{pl_last}")
    _cf(pl, f"B2:B{pl_last}", 'B2="Issued"', "DBEAFE"); _cf(pl, f"B2:B{pl_last}", 'B2="In Field"', "E0E7FF")
    _cf(pl, f"B2:B{pl_last}", 'B2="Returned"', "FEF3C7"); _cf(pl, f"B2:B{pl_last}", 'B2="Filed"', "DCFCE7")
    _cf(pl, f"B2:B{pl_last}", 'B2="Rejected"', "FECACA")

    # ---- Signature Sheets -----------------------------------------------------------
    ss = _sheet(wb, "Signature Sheets", COLS["Signature Sheets"],
                {"A": 12, "B": 8, "C": 12, "D": 22, "E": 13, "F": 13, "G": 13, "H": 14, "I": 12, "J": 12, "K": 10, "L": 12, "M": 32, "N": 22, "O": 16, "P": 14, "Q": 14})
    if template_mode:
        sh_rows = [(f"P-{i:03d}", k, None, None, None, None, "Blank", 0, 0, 0, None, None, None, None, None)
                   for i in range(1, run + 1) for k in range(1, spp + 1)]
    else:
        sh_rows = [(p.number, sh.sheet_no, sh.circulator.name if sh.circulator else (p.issued_to.name if p.issued_to else None),
                    sh.issued_on, sh.returned_on, sh.notarized_on, sh.status, sh.collected, sh.questionable, sh.rejected, sh.notes,
                    sh.notary_name, sh.notary_commission, sh.notary_expiration, sh.defect_codes)
                   for p in pamphlets for sh in p.sheets]
    for r, (num, k, circ, iss, ret, nd, st, col, qu, rj, notes, nn, nc, ne, dc) in enumerate(sh_rows, 2):
        ss.cell(r, 1, num); ss.cell(r, 2, k); ss.cell(r, 3, f"{num}-S{k}"); ss.cell(r, 4, circ); ss.cell(r, 5, iss); ss.cell(r, 6, ret)
        ss.cell(r, 7, nd); ss.cell(r, 8, st); ss.cell(r, 9, col); ss.cell(r, 10, qu); ss.cell(r, 11, rj)
        ss.cell(r, 12, f"=MAX(I{r}-J{r}-K{r},0)").fill = CALC_FILL
        ss.cell(r, 13, notes); ss.cell(r, 14, nn); ss.cell(r, 15, nc); ss.cell(r, 16, ne); ss.cell(r, 17, dc)
        for c in (5, 6, 7, 16): ss.cell(r, c).number_format = DATE_FMT
    ss_last = max(len(sh_rows) + 1, 2)
    _table(ss, "Table_1", f"A1:Q{ss_last}")
    _dv(ss, dd_ref["Sheet Status"], f"H2:H{ss_last}")
    _dv(ss, "=Volunteers!$A$2:$A$201", f"D2:D{ss_last}")
    _cf(ss, f"H2:H{ss_last}", 'H2="Needs Fix"', "FEE2E2"); _cf(ss, f"H2:H{ss_last}", 'H2="Rejected"', "FECACA")
    _cf(ss, f"H2:H{ss_last}", 'H2="Audited OK"', "DCFCE7"); _cf(ss, f"H2:H{ss_last}", 'H2="Notarized"', "DBEAFE")
    _cf(ss, f"I2:I{ss_last}", f"I2>Dashboard!$B$5", "FECACA")   # more signatures than lines on the sheet

    # ---- Dashboard (inserted at index 1 so it opens first) --------------------------
    d = wb.create_sheet("Dashboard", 1)
    d["A1"] = f"Petition Captain Dashboard — {run} Pamphlet Print Run"; d["A1"].font = Font(size=16, bold=True); d.merge_cells("A1:H1")
    d["A2"] = "Yellow cells are inputs. Everything else is computed. The admin area at petition.mcalester.net is the system of record."; d["A2"].font = Font(italic=True, color="555555"); d.merge_cells("A2:H2")
    inputs = [
        (3, "Print Run", run, None),
        (4, "Signature Sheets per Pamphlet", spp, None),
        (5, "Signature Lines per Sheet", rps, None),
        (6, "Total Signature Capacity", "=B3*B4*B5", None),
        (7, "Legal Minimum Needed (10% of registered voters — 62 O.S. § 868(B)(2))", '=IF(B13="","",ROUNDUP(B13*B14,0))', None),
        (8, "Overcollection Target %", s.float("overcollect_fraction") or 0.5, "0%"),
        (9, "Target Signatures", '=IF(B7="","",ROUNDUP(B7*(1+B8),0))', None),
        (10, "Signatures Collected (from Signature Sheets)", f"=SUM('Signature Sheets'!I2:I{ss_last})", None),
        (11, "Remaining to Target", '=IF(B9="","",MAX(B9-B10,0))', None),
        (12, "Estimated Valid %", s.est_valid_rate, "0%"),
        (13, "Registered Voters (written count from County Election Board)", s.registered_voters, "#,##0"),
        (14, "Legal Fraction (62 O.S. § 868(B)(2))", petition.threshold.legal_fraction, "0%"),
        (15, "Registered Voter Count — Source", s.raw("registered_voters_source"), None),
        (16, "Registered Voter Count — Date", s.date("registered_voters_date"), DATE_FMT),
        (17, "Adoption Date (Board vote — TABLED, blank until adopted)", s.adoption_date, DATE_FMT),
        (18, "Filing Deadline (adoption + 30 days — 62 O.S. § 868(B)(3))", '=IF(B17="","",B17+30)', DATE_FMT),
        (19, "Days Remaining to File", '=IF(B18="","",B18-TODAY())', "0"),
        (20, "Election Date (next general county election — 62 O.S. § 868(H))", s.election_date, DATE_FMT),
    ]
    for row, label, val, fmt in inputs:
        d.cell(row, 1, label).font = Font(bold=True)
        c = d.cell(row, 2, val)
        if fmt: c.number_format = fmt
        c.fill = CALC_FILL if isinstance(val, str) and val.startswith("=") else INPUT_FILL
        c.border = BOX
    glance = [
        (3, "At-a-Glance", None, None, None),
        (4, "Pamphlets Printed", f"=COUNTIF('Pamphlet Log'!B2:B{pl_last},\"Printed\")+COUNTIF('Pamphlet Log'!B2:B{pl_last},\"Issued\")+COUNTIF('Pamphlet Log'!B2:B{pl_last},\"In Field\")+COUNTIF('Pamphlet Log'!B2:B{pl_last},\"Returned\")+COUNTIF('Pamphlet Log'!B2:B{pl_last},\"Audited\")+COUNTIF('Pamphlet Log'!B2:B{pl_last},\"Filed\")",
         "Sheets Notarized", f"=COUNTIF('Signature Sheets'!H2:H{ss_last},\"Notarized\")+COUNTIF('Signature Sheets'!H2:H{ss_last},\"Audited OK\")+COUNTIF('Signature Sheets'!H2:H{ss_last},\"Filed\")"),
        (5, "Pamphlets Issued", f"=COUNTIF('Pamphlet Log'!B2:B{pl_last},\"Issued\")+COUNTIF('Pamphlet Log'!B2:B{pl_last},\"In Field\")+COUNTIF('Pamphlet Log'!B2:B{pl_last},\"Returned\")+COUNTIF('Pamphlet Log'!B2:B{pl_last},\"Audited\")+COUNTIF('Pamphlet Log'!B2:B{pl_last},\"Filed\")",
         "Audited OK Sheets", f"=COUNTIF('Signature Sheets'!H2:H{ss_last},\"Audited OK\")+COUNTIF('Signature Sheets'!H2:H{ss_last},\"Filed\")"),
        (6, "Pamphlets Returned", f"=COUNTIF('Pamphlet Log'!B2:B{pl_last},\"Returned\")+COUNTIF('Pamphlet Log'!B2:B{pl_last},\"Audited\")+COUNTIF('Pamphlet Log'!B2:B{pl_last},\"Filed\")",
         "Rejected Sheets", f"=COUNTIF('Signature Sheets'!H2:H{ss_last},\"Rejected\")"),
        (7, "Pamphlets Filed", f"=COUNTIF('Pamphlet Log'!B2:B{pl_last},\"Filed\")",
         "Open Issues", "=COUNTIF(Issues!F:F,\"Open\")+COUNTIF(Issues!F:F,\"Investigating\")+COUNTIF(Issues!F:F,\"Escalated\")"),
        (8, "Total Capacity", "=B6", "Est. Valid Signatures", "=ROUND(B10*B12,0)"),
        (9, "Collection Progress (to target)", '=IF(B9="","",B10/B9)', "Capacity Used", '=IF(B6=0,"",B10/B6)'),
        (10, "Valid Estimate (sheet audit)", f"=SUM('Signature Sheets'!L2:L{ss_last})", "Progress to Legal Minimum", '=IF(B7="","",E10/B7)'),
        (11, "DB cross-check: collected", stats["collected"], "DB cross-check: valid estimate", stats["valid_estimate"]),
        (12, "DB cross-check: pamphlets filed", stats["pamphlets"]["Filed"], "DB cross-check: open issues", stats["open_issues"]),
        (13, "Exported", now.strftime("%Y-%m-%d %H:%M"), "Circulators cleared to circulate (DB)", stats["circulators_ready"]),
    ]
    for row, l1, v1, l2, v2 in glance:
        d.cell(row, 4, l1).font = Font(bold=True)
        if v1 is not None:
            c = d.cell(row, 5, v1); c.border = BOX; c.fill = CALC_FILL
        if l2: d.cell(row, 7, l2).font = Font(bold=True)
        if v2 is not None:
            c = d.cell(row, 8, v2); c.border = BOX; c.fill = CALC_FILL
    for cell in ("E9", "H9", "H10"): d[cell].number_format = "0%"
    d["D3"].font = Font(size=12, bold=True); d.merge_cells("D3:H3")
    rules = ["Hard Rules for Petition Captains",
             "No signatures until the true copy is filed with the Secretary of the County Election Board (62 O.S. § 868(B)(1)).",
             "Use only the final, filed, identical pamphlet version — check the version hash on every print batch.",
             "Print and circulate complete pamphlets only, never loose signature sheets (34 O.S. § 3; § 6.1(A)(3)).",
             "Each affidavit stays directly behind the signature sheet it verifies (34 O.S. § 6).",
             "Every circulator is a registered Oklahoma voter, personally witnesses every signature, and signs the affidavit after collection, before a notary.",
             "Do not detach, reorder, replace, or photocopy individual pages. Do not photograph signed sheets.",
             f"Track all {run} pamphlets by pamphlet number and assigned volunteer. Return packets daily.",
             "Nothing in this workbook or the website ever records a signer's name, address or birth date."]
    for i, txt in enumerate(rules, 23):
        c = d.cell(i, 1, txt); c.font = Font(bold=(i == 23), size=12 if i == 23 else 11); d.merge_cells(f"A{i}:H{i}")
    _widths(d, {"A": 58, "B": 16, "C": 3, "D": 34, "E": 14, "F": 3, "G": 36, "H": 14})
    _print_setup(d, title_rows=None)
    for name, ref in (("RegisteredVoters", "Dashboard!$B$13"), ("AdoptionDate", "Dashboard!$B$17"), ("FilingDeadline", "Dashboard!$B$18"),
                      ("LegalMinimum", "Dashboard!$B$7"), ("TargetSignatures", "Dashboard!$B$9"), ("ElectionDate", "Dashboard!$B$20")):
        wb.defined_names[name] = DefinedName(name, attr_text=ref)

    # ---- Daily Counts (day 0 = adoption day … day 30 = deadline) --------------------
    dc = _sheet(wb, "Daily Counts", COLS["Daily Counts"], {"A": 13, "B": 12, "C": 12, "D": 10, "E": 12, "F": 12, "G": 28, "H": 28})
    for i in range(0, 31):
        r = i + 2
        dc.cell(r, 1, f'=IF(AdoptionDate="","",AdoptionDate+{i})').number_format = DATE_FMT
        dc.cell(r, 2, f"=IF(A{r}=\"\",0,SUMIFS('Signature Sheets'!$I:$I,'Signature Sheets'!$F:$F,A{r}))")
        dc.cell(r, 3, f"=IF(A{r}=\"\",0,SUMIFS('Signature Sheets'!$J:$J,'Signature Sheets'!$F:$F,A{r}))")
        dc.cell(r, 4, f"=IF(A{r}=\"\",0,SUMIFS('Signature Sheets'!$K:$K,'Signature Sheets'!$F:$F,A{r}))")
        dc.cell(r, 5, f"=MAX(B{r}-C{r}-D{r},0)")
        dc.cell(r, 6, f"=IF(A{r}=\"\",0,COUNTIFS('Pamphlet Log'!$F:$F,A{r}))")
        for c in range(1, 7): dc.cell(r, c).fill = CALC_FILL
    dc.cell(33, 1, "Days count from the Adoption Date on the Dashboard (day 0) to the 30-day filing deadline. Counts roll up from Signature Sheets by Returned Date.").font = Font(italic=True)
    _table(dc, "Table_3", "A1:H32")

    # ---- Volunteers -----------------------------------------------------------------
    vs = _sheet(wb, "Volunteers", COLS["Volunteers"], {"A": 24, "B": 18, "C": 16, "D": 26, "E": 14, "F": 14, "G": 10, "H": 12, "I": 22, "J": 30, "K": 12, "L": 16, "M": 12, "N": 13, "O": 9})
    vols = db.scalars(select(m.Circulator).order_by(m.Circulator.name)).all()
    for r, v in enumerate(vols, 2):
        vals = [v.name, v.role, v.phone, v.email, _yes(v.registered_voter_verified), _yes(v.trained_on), _yes(v.is_notary), None,
                v.availability, v.notes, v.registered_verified_on, v.registered_verified_by, v.trained_on, _yes(v.compensated), _yes(v.active)]
        for c, val in enumerate(vals, 1): vs.cell(r, c, val)
    v_last = max(len(vols) + 1, 2)
    for r in range(2, max(v_last, 200) + 1):
        vs.cell(r, 8, f"=IF(A{r}=\"\",\"\",COUNTIF('Pamphlet Log'!$D:$D,A{r}))").fill = CALC_FILL
        for c in (11, 13): vs.cell(r, c).number_format = DATE_FMT
    _table(vs, "Table_4", f"A1:O{max(v_last, 200)}")
    _dv(vs, dd_ref["Volunteer Role"], "B2:B200")
    for col in "EFGNO": _dv(vs, dd_ref["Yes/No"], f"{col}2:{col}200")
    _cf(vs, "E2:E200", 'AND(A2<>"",E2<>"Yes")', "FEE2E2")   # a circulator not yet verified as a registered voter

    # ---- Filing QA ------------------------------------------------------------------
    qa = _sheet(wb, "Filing QA", COLS["Filing QA"], {"A": 90, "B": 16, "C": 20, "D": 50})
    tasks = db.scalars(select(m.QATask).order_by(m.QATask.sort_order, m.QATask.id)).all()
    qa_rows = [(t.task, t.status, t.owner, t.notes) for t in tasks] or [(t.format(run=run), "Not Started", None, None) for t in DEFAULT_QA_TASKS]
    for r, row in enumerate(qa_rows, 2):
        for c, val in enumerate(row, 1): qa.cell(r, c, val)
    _table(qa, "Table_5", f"A1:D{len(qa_rows) + 1}")
    _dv(qa, dd_ref["QA Status"], f"B2:B{len(qa_rows) + 30}")
    _cf(qa, f"B2:B{len(qa_rows) + 30}", 'B2="Done"', "DCFCE7"); _cf(qa, f"B2:B{len(qa_rows) + 30}", 'B2="Blocked"', "FECACA")

    # ---- EventsShifts ---------------------------------------------------------------
    ev = _sheet(wb, "EventsShifts", COLS["EventsShifts"], {"A": 13, "B": 8, "C": 8, "D": 34, "E": 22, "F": 12, "G": 12, "H": 14, "I": 32})
    events = db.scalars(select(m.Event).order_by(m.Event.date, m.Event.start)).all()
    for r, e in enumerate(events, 2):
        vals = [e.date, e.start.strftime("%H:%M") if e.start else None, e.end.strftime("%H:%M") if e.end else None,
                e.location.name if e.location else None, e.lead.name if e.lead else None, e.volunteers_needed, e.pamphlets_issued, e.expected_signatures, e.notes]
        for c, val in enumerate(vals, 1): ev.cell(r, c, val)
        ev.cell(r, 1).number_format = DATE_FMT
    _table(ev, "Table_6", f"A1:I{max(len(events) + 1, 100)}")

    # ---- Issues ---------------------------------------------------------------------
    iss = _sheet(wb, "Issues", COLS["Issues"], {"A": 10, "B": 13, "C": 12, "D": 12, "E": 16, "F": 14, "G": 10, "H": 50})
    issues = db.scalars(select(m.Issue).order_by(m.Issue.number)).all()
    i_rows = [(i.number, i.opened_on, i.pamphlet.number if i.pamphlet else None, i.sheet.sheet_id if i.sheet else None,
               i.issue_type, i.status, i.priority, i.notes) for i in issues]
    n_issue_rows = max(len(i_rows), 300)
    for r in range(2, n_issue_rows + 2):
        k = r - 2
        row = i_rows[k] if k < len(i_rows) else (f"I-{k + 1:03d}", None, None, None, None, None, None, None)
        for c, val in enumerate(row, 1): iss.cell(r, c, val)
        iss.cell(r, 2).number_format = DATE_FMT
    i_last = n_issue_rows + 1
    _table(iss, "Table_7", f"A1:H{i_last}")
    _dv(iss, dd_ref["Issue Status"], f"F2:F{i_last}"); _dv(iss, dd_ref["Priority"], f"G2:G{i_last}"); _dv(iss, dd_ref["Issue Type"], f"E2:E{i_last}")
    _dv(iss, "='Pamphlet Log'!$A$2:$A$" + str(pl_last), f"C2:C{i_last}")
    _cf(iss, f"F2:F{i_last}", 'F2="Open"', "FEF3C7"); _cf(iss, f"G2:G{i_last}", 'G2="Critical"', "FECACA")

    # ---- Deadlines ------------------------------------------------------------------
    dl = _sheet(wb, "Deadlines", COLS["Deadlines"], {"A": 46, "B": 22, "C": 60, "D": 14, "E": 40, "F": 30})
    dl_rows = [
        ("Board of County Commissioners adopts the resolution", "62 O.S. § 868(B)(3)", "Starts the 30-day referendum clock. TABLED — no date yet.", "=IF(AdoptionDate=\"\",\"\",AdoptionDate)", "Dashboard → Adoption Date", None),
        ("True copy of the petition filed with the Secretary of the County Election Board", "62 O.S. § 868(B)(1)", "Must happen BEFORE any signature is collected. Get a file-stamped copy.", None, "ENTER the filing date (yellow)", None),
        ("Ballot title filed with the Secretary", "62 O.S. § 868(D)(1)", "≤150 words, gist, yes/no language, no partiality. File it with the true copy.", None, "ENTER the filing date (yellow)", None),
        ("District Attorney's review of the ballot title complete", "62 O.S. § 868(D)(2)", "Within 3 days after the ballot title is filed the DA says whether it is in legal form, or files a conforming one.", "=IF(D4=\"\",\"\",D4+3)", "= ballot title filed + 3 days", None),
        ("Ballot-title appeal window closes", "62 O.S. § 868(E)", "Any qualified elector may appeal the wording to district court within 10 days after the ballot title is filed.", "=IF(D4=\"\",\"\",D4+10)", "= ballot title filed + 10 days", None),
        ("SIGNED PETITIONS DUE to the Secretary of the County Election Board", "62 O.S. § 868(B)(3)", "Within 30 days after adoption. File earlier. Bring every pamphlet, organized and logged; get a receipt.", "=IF(FilingDeadline=\"\",\"\",FilingDeadline)", "= adoption + 30 days (Dashboard)", None),
        ("County publishes notice of filing and apparent sufficiency", "62 O.S. § 868(C)", "Secretary counts signatures and publishes in a newspaper of general circulation.", None, "ENTER the publication date (yellow)", None),
        ("Protest / objection window closes", "62 O.S. § 868(C)", "Protest to the petition or the count is filed in district court within 10 days after publication.", "=IF(D8=\"\",\"\",D8+10)", "= publication + 10 days", None),
        ("Election", "62 O.S. § 868(H)", "The question goes to county voters at the next general county election.", "=IF(ElectionDate=\"\",\"\",ElectionDate)", "Dashboard → Election Date", None),
    ]
    for r, row in enumerate(dl_rows, 2):
        for c, val in enumerate(row, 1):
            cell = dl.cell(r, c, val); cell.alignment = Alignment(wrap_text=True, vertical="top")
        dcell = dl.cell(r, 4); dcell.number_format = DATE_FMT; dcell.border = BOX
        dcell.fill = CALC_FILL if row[3] else INPUT_FILL
    _table(dl, "Table_Deadlines", f"A1:F{len(dl_rows) + 1}")

    # ---- Notary Log -----------------------------------------------------------------
    nl = _sheet(wb, "Notary Log", COLS["Notary Log"], {"A": 13, "B": 12, "C": 8, "D": 22, "E": 10, "F": 22, "G": 14, "H": 13, "I": 10, "J": 30})
    n_rows = [(sh.notarized_on, p.number, sh.sheet_no, sh.circulator.name if sh.circulator else (p.issued_to.name if p.issued_to else None),
               sh.collected, sh.notary_name, sh.notary_commission, sh.notary_expiration, None, sh.notes)
              for p in pamphlets for sh in p.sheets if sh.notarized_on or sh.notary_name]
    for r, row in enumerate(n_rows, 2):
        for c, val in enumerate(row, 1): nl.cell(r, c, val)
        nl.cell(r, 1).number_format = DATE_FMT; nl.cell(r, 8).number_format = DATE_FMT
    _table(nl, "Table_Notary", f"A1:J{max(len(n_rows) + 1, 40)}")
    nl.cell(max(len(n_rows) + 1, 40) + 2, 1, "One line per notarized affidavit. Check before leaving the notary: signature, printed name/title, address, commission number, expiration date, seal (34 O.S. § 6; § 6.1(A)(7)).").font = Font(italic=True)

    # ---- Records Log ----------------------------------------------------------------
    rl = _sheet(wb, "Records Log", COLS["Records Log"], {"A": 18, "B": 60, "C": 30, "D": 22, "E": 40, "F": 12, "G": 36})
    recs = db.scalars(select(m.RecordsLog).order_by(m.RecordsLog.occurred_at, m.RecordsLog.id)).all()
    r_rows = [(x.occurred_at.replace(tzinfo=None) if x.occurred_at else None, x.item, x.office, x.person, x.documents, _yes(x.receipt_obtained), x.notes) for x in recs] \
        or [(None, it, None, None, None, None, None) for it in DEFAULT_RECORDS]
    for r, row in enumerate(r_rows, 2):
        for c, val in enumerate(row, 1): rl.cell(r, c, val)
        rl.cell(r, 1).number_format = "yyyy-mm-dd hh:mm"
    _table(rl, "Table_Records", f"A1:G{max(len(r_rows) + 1, 30)}")
    _dv(rl, dd_ref["Yes/No"], "F2:F200")

    # ---- Contacts -------------------------------------------------------------------
    ct = _sheet(wb, "Contacts", COLS["Contacts"], {"A": 40, "B": 26, "C": 16, "D": 32, "E": 44, "F": 26, "G": 9})
    contacts = db.scalars(select(m.Contact).order_by(m.Contact.sort_order, m.Contact.id)).all()
    c_rows = [(x.role, x.name, x.phone, x.email, x.address, x.hours, _yes(x.public)) for x in contacts]
    if not c_rows:
        eb = petition.contacts.get("election_board", {})
        c_rows = [("Petition Captain", s.raw("captain_name") or "[NAME]", s.raw("captain_phone") or "[PHONE]", None, None, None, "Yes"),
                  (f"{eb.get('name', 'County Election Board')} — {eb.get('secretary', '')}, Secretary", eb.get("secretary"), eb.get("phone"), eb.get("email"), eb.get("address"), eb.get("hours"), "Yes"),
                  ("Election-law attorney", None, None, None, None, None, "No"), ("District Attorney (ballot-title review, 62 O.S. § 868(D)(2))", None, None, None, None, None, "No"),
                  ("Print vendor", None, None, None, None, None, "No")]
    for r, row in enumerate(c_rows, 2):
        for c, val in enumerate(row, 1): ct.cell(r, c, val)
    _table(ct, "Table_Contacts", f"A1:G{max(len(c_rows) + 1, 20)}")
    _dv(ct, dd_ref["Yes/No"], "G2:G100")

    # ---- Precincts ------------------------------------------------------------------
    pc = _sheet(wb, "Precincts", COLS["Precincts"], {"A": 10, "B": 44, "C": 30, "D": 14, "E": 14, "F": 18, "G": 18, "H": 12, "I": 12})
    prs = _precincts()
    for r, p in enumerate(prs, 2):
        for c, val in enumerate([p["precinct"], p["polling_place"], p["address"], p["city"], p["pop"], p["vap"], p["comm"], p["house"], p["senate"]], 1):
            pc.cell(r, c, val)
    pc.cell(len(prs) + 3, 1, "Source: OU Center for Spatial Analysis precinct layer (State Election Board GIS contractor) + Pittsburg County Election Board polling places. 2020 Census counts are population, NOT registered voters.").font = Font(italic=True)
    _table(pc, "Table_Precincts", f"A1:I{len(prs) + 1}")

    # ---- Defect Codes ---------------------------------------------------------------
    df = _sheet(wb, "Defect Codes", COLS["Defect Codes"], {"A": 8, "B": 90, "C": 22, "D": 80})
    for r, (code, txt) in enumerate(zip(m.DEFECT_CODES, statutes.exclusions()), 2):
        for c, val in enumerate([code, txt, DEFECT_SCOPE[code], DEFECT_PREVENT[code]], 1):
            cell = df.cell(r, c, val); cell.alignment = Alignment(wrap_text=True, vertical="top")
        if DEFECT_SCOPE[code].startswith("WHOLE"): df.cell(r, 3).font = Font(bold=True, color="B91C1C")
    df.cell(11, 1, f"Verbatim from {statutes.cite('34-6.1')} — {statutes.cite_url('34-6.1')} (retrieved {statutes.header('34-6.1').get('retrieved', '')}). Use these codes in the Defect Codes column of Signature Sheets and the Issue Type column of Issues.").font = Font(italic=True)
    _table(df, "Table_Defects", "A1:D9")

    # ---- RulesSources ---------------------------------------------------------------
    rs = _sheet(wb, "RulesSources", COLS["RulesSources"], {"A": 34, "B": 16, "C": 44, "D": 90, "E": 60, "F": 12})
    secs = list(petition.statutes)
    for r, sec in enumerate(secs, 2):
        h = statutes.header(sec)
        for c, val in enumerate([h.get("title", sec).split(" — ", 1)[-1], statutes.cite(sec), h.get("title", ""), STATUTE_MEANING.get(sec, ""), statutes.cite_url(sec), h.get("retrieved", "")], 1):
            cell = rs.cell(r, c, val); cell.alignment = Alignment(wrap_text=True, vertical="top")
    _table(rs, "Table_8", f"A1:F{len(secs) + 1}")

    # ---- final ordering: Dashboard first; README right after; Dropdowns last -------
    order = ["Dashboard", "README", "Pamphlet Log", "Signature Sheets", "Daily Counts", "Volunteers", "Filing QA", "EventsShifts", "Issues",
             "Deadlines", "Notary Log", "Records Log", "Contacts", "Precincts", "Defect Codes", "RulesSources", "Dropdowns"]
    wb._sheets = [wb[n] for n in order]
    wb.active = 0
    return wb


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--out", default="output/xlsx/petition-master.xlsx")
    ap.add_argument("--database-url", default=None, help="defaults to $DATABASE_URL or the local SQLite dev database")
    a = ap.parse_args(argv)
    from app.db import make_engine, init_db
    from sqlalchemy.orm import sessionmaker
    eng = make_engine(a.database_url)
    init_db(eng)
    with sessionmaker(bind=eng, expire_on_commit=False)() as db:
        wb = build_workbook(db)
    out = Path(a.out); out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out} ({', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    main()
