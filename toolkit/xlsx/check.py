"""Structural checks on an exported Petition Master workbook (no recalculation engine here).

    python -m toolkit.xlsx.check output/xlsx/petition-master.xlsx
"""
from __future__ import annotations
import argparse, re, sys
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

REQUIRED = {
    "Dashboard": [], "README": [],
    "Pamphlet Log": ["Pamphlet #", "Status", "Printed Date", "Issued To", "Issued Date", "Returned Date", "Sheet Count", "Expected Capacity",
                     "Collected Count", "Notarized Sheets", "Audited OK Sheets", "Rejected Sheets", "Filed?", "Notes"],
    "Signature Sheets": ["Pamphlet #", "Sheet #", "Sheet ID", "Circulator", "Issued Date", "Returned Date", "Notarized Date", "Status",
                         "Collected Signatures", "Questionable", "Rejected", "Valid Estimate", "Notes", "Defect Codes"],
    "Daily Counts": ["Date", "Collected Signatures", "Questionable", "Rejected", "Valid Estimate", "Pamphlets Returned"],
    "Volunteers": ["Name", "Role", "Phone", "Email", "County Registered Voter?", "Training Complete?", "Notary?", "Assigned Pamphlets"],
    "Filing QA": ["Task", "Status", "Owner", "Notes"],
    "EventsShifts": ["Date", "Start", "End", "Location", "Event Lead"],
    "Issues": ["Issue #", "Date", "Pamphlet #", "Sheet ID", "Issue Type", "Status", "Priority"],
    "Deadlines": ["Step", "Statute", "Rule", "Date"], "Notary Log": ["Date", "Pamphlet #", "Sheet #", "Notary Name", "Commission #", "Expiration"],
    "Records Log": ["Date / Time", "Item", "Office", "Person"], "Contacts": ["Role", "Name", "Phone"],
    "Precincts": ["Precinct", "Polling Place", "2020 Population"], "Defect Codes": ["Code"], "RulesSources": ["Cite", "Source URL"], "Dropdowns": [],
}
REQUIRED_DV = {"Pamphlet Log": "B", "Signature Sheets": "H", "Volunteers": "B", "Issues": "F", "Filing QA": "B"}
FORBIDDEN_STRINGS = ["June 22", "July 22", "2759"]
PII_HEADERS = re.compile(r"^(signature|signer|voter name)$|birth|first name|last name|street|house #|date of birth|\bdob\b", re.I)
REF_RE = re.compile(r"(?:'((?:[^']|'')+)'|([A-Za-z_][\w.]*))!(\$?[A-Z]{1,3})(\$?\d+)?(?::(\$?[A-Z]{1,3})(\$?\d+)?)?")


def check(path: str) -> list[str]:
    errs: list[str] = []
    wb = load_workbook(path)
    names = set(wb.sheetnames)
    defined = set(wb.defined_names.keys())
    for sheet, cols in REQUIRED.items():
        if sheet not in names:
            errs.append(f"missing sheet {sheet!r}"); continue
        hdr = [str(c.value).strip() if c.value is not None else "" for c in wb[sheet][1]]
        for col in cols:
            if col not in hdr: errs.append(f"{sheet}: missing column {col!r}")
        for h in hdr:
            if h and PII_HEADERS.search(h): errs.append(f"{sheet}: PII-looking column header {h!r}")
    if "Dashboard" in names:
        dash = wb["Dashboard"]
        labels = {str(dash.cell(r, 1).value or "").lower(): dash.cell(r, 2).value for r in range(1, 40)}
        lm = next((v for k, v in labels.items() if k.startswith("legal minimum needed")), None)
        if not (isinstance(lm, str) and lm.startswith("=")): errs.append("Dashboard: Legal Minimum must be a formula from the registered-voter cell, not a typed number")
        if not any(k.startswith("adoption date") for k in labels): errs.append("Dashboard: Adoption Date input row missing")
    for name in ("AdoptionDate", "FilingDeadline", "RegisteredVoters"):
        if name not in defined: errs.append(f"defined name {name} missing")

    def check_ref(where, formula):
        for q, u, c1, r1, c2, r2 in REF_RE.findall(formula):
            sheet = (q or u).replace("''", "'")
            if sheet not in names:
                errs.append(f"{where}: formula references unknown sheet {sheet!r}: {formula[:80]}"); continue
            ws = wb[sheet]
            for col, row in ((c1, r1), (c2, r2)):
                if not col: continue
                ci = column_index_from_string(col.replace("$", ""))
                if ci > max(ws.max_column, 1): errs.append(f"{where}: {sheet}!{col} beyond last column ({ws.max_column})")
                if row and int(row.replace("$", "")) > max(ws.max_row, 1) + 1000: errs.append(f"{where}: {sheet}!{col}{row} far beyond last row ({ws.max_row})")

    for ws in wb.worksheets:
        if ws.page_setup.orientation != "landscape" or ws.page_setup.paperSize != 5:
            errs.append(f"{ws.title}: print setup must be legal landscape (paperSize=5)")
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str):
                    if v.startswith("="): check_ref(f"{ws.title}!{c.coordinate}", v)
                    for bad in FORBIDDEN_STRINGS:
                        if bad in v: errs.append(f"{ws.title}!{c.coordinate}: contains forbidden {bad!r}")
                elif isinstance(v, (int, float)) and v == 2759:
                    errs.append(f"{ws.title}!{c.coordinate}: contains the stale 2759 figure")
        for dv in ws.data_validations.dataValidation:
            if dv.formula1 and "!" in dv.formula1: check_ref(f"{ws.title} DV", dv.formula1)
        for name, col in REQUIRED_DV.items():
            if ws.title == name and not any(col in str(dv.sqref) for dv in ws.data_validations.dataValidation):
                errs.append(f"{name}: dropdown validation missing on column {col}")
    for name in defined:
        dn = wb.defined_names[name]
        m = REF_RE.match(dn.attr_text or "")
        if not m or (m.group(1) or m.group(2)) not in names: errs.append(f"defined name {name} -> {dn.attr_text!r} does not resolve")
    return errs


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("path")
    a = ap.parse_args(argv)
    errs = check(a.path)
    if errs:
        print(f"FAIL {a.path}: {len(errs)} problem(s)")
        for e in errs: print("  -", e)
        sys.exit(1)
    print(f"OK {a.path}")


if __name__ == "__main__":
    main()
