"""One-time (idempotent) import of the legacy "Petition Captain Master Tracker.xlsx" — or a
previous export of ours — into the database.

    python -m toolkit.xlsx.import_tracker "Petition Captain Master Tracker.xlsx" [--database-url URL]

Upserts by natural keys (pamphlet number, (pamphlet, sheet #), issue number, volunteer name,
QA task text, location name, (location, date, start)). Statuses map 1:1. The tracker's
hard-coded "Legal Minimum Needed" is NOT imported as a registered-voter count: that figure
stays a placeholder until the County Election Board provides a written count.
"""
from __future__ import annotations
import argparse, logging, re
from datetime import date, datetime, time
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session
from app import models as m
from app.settings import Settings

log = logging.getLogger("xlsx.import")
# Legacy task wording that baked in the (now tabled) June 22 vote date.
LEGACY_TEXT_FIXES = [
    (re.compile(r"after (the )?June 22(,? 2026)? vote", re.I), "after the Board adopts the resolution (tabled — date TBD)"),
    (re.compile(r"\bJune 22(,? 2026)?\b"), "the adoption date (TBD)"),
    (re.compile(r"\bJuly 22(,? 2026)?\b"), "the 30-day deadline (adoption + 30)"),
]


def _fix_text(s: str | None) -> str | None:
    if not s: return s
    for rx, rep in LEGACY_TEXT_FIXES: s = rx.sub(rep, s)
    return s


def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")[:64]


def _date(v) -> date | None:
    if v in (None, ""): return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    try: return date.fromisoformat(str(v)[:10])
    except ValueError: return None


def _time(v) -> time | None:
    if v in (None, ""): return None
    if isinstance(v, datetime): return v.time()
    if isinstance(v, time): return v
    try: return time.fromisoformat(str(v))
    except ValueError: return None


def _int(v) -> int | None:
    if v in (None, ""): return None
    try: return int(float(v))
    except (TypeError, ValueError): return None


def _yes(v) -> bool:
    return str(v).strip().lower() in ("yes", "y", "true", "1", "x")


def _str(v) -> str | None:
    if v is None: return None
    s = str(v).strip()
    return s or None


def _rows(ws, headers_wanted: list[str]) -> list[dict]:
    """Rows as dicts keyed by header text (case-insensitive, trimmed)."""
    hdr = {str(c.value).strip().lower(): i for i, c in enumerate(ws[1]) if c.value is not None}
    missing = [h for h in headers_wanted if h.lower() not in hdr]
    if missing:
        raise ValueError(f"{ws.title}: missing columns {missing}")
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in r): continue
        out.append({h: (r[i] if i < len(r) else None) for h, i in hdr.items()})
    return out


def _get(row: dict, *names):
    for n in names:
        v = row.get(n.lower())
        if v not in (None, ""): return v
    return None


def import_tracker(path: str | Path, db: Session) -> dict:
    wb = load_workbook(path, data_only=True)
    counts = {"pamphlets": 0, "sheets": 0, "circulators": 0, "issues": 0, "qa_tasks": 0, "locations": 0, "events": 0, "settings": 0, "warnings": []}
    warn = counts["warnings"]

    def circulator(name) -> m.Circulator | None:
        name = _str(name)
        if not name: return None
        c = db.scalar(select(m.Circulator).where(m.Circulator.name == name))
        if c is None:
            c = m.Circulator(name=name, role="Circulator"); db.add(c); db.flush(); counts["circulators"] += 1
        return c

    # ---- Volunteers first so pamphlet "Issued To" can resolve --------------------
    if "Volunteers" in wb.sheetnames:
        for row in _rows(wb["Volunteers"], ["Name"]):
            name = _str(_get(row, "Name"))
            if not name: continue
            c = db.scalar(select(m.Circulator).where(m.Circulator.name == name))
            if c is None:
                c = m.Circulator(name=name); db.add(c); counts["circulators"] += 1
            c.role = _str(_get(row, "Role")) or c.role or "Circulator"
            c.phone = _str(_get(row, "Phone")) or c.phone
            c.email = _str(_get(row, "Email")) or c.email
            c.registered_voter_verified = _yes(_get(row, "County Registered Voter?"))
            c.registered_verified_on = _date(_get(row, "Verified On")) or c.registered_verified_on
            c.registered_verified_by = _str(_get(row, "Verified By")) or c.registered_verified_by
            if _yes(_get(row, "Training Complete?")):
                c.trained_on = _date(_get(row, "Trained On")) or c.trained_on or date.today()
            else:
                c.trained_on = _date(_get(row, "Trained On"))
            c.is_notary = _yes(_get(row, "Notary?"))
            c.compensated = _yes(_get(row, "Compensated?"))
            c.availability = _str(_get(row, "Availability")) or c.availability
            c.notes = _str(_get(row, "Notes")) or c.notes
            if _get(row, "Active?") is not None: c.active = _yes(_get(row, "Active?"))
        db.flush()

    # ---- Pamphlet Log ----------------------------------------------------------------
    pam_by_num: dict[str, m.Pamphlet] = {}
    if "Pamphlet Log" in wb.sheetnames:
        for row in _rows(wb["Pamphlet Log"], ["Pamphlet #", "Status"]):
            num = _str(_get(row, "Pamphlet #"))
            if not num: continue
            p = db.scalar(select(m.Pamphlet).where(m.Pamphlet.number == num))
            if p is None:
                p = m.Pamphlet(number=num); db.add(p); counts["pamphlets"] += 1
            st = _str(_get(row, "Status"))
            if st in m.PAMPHLET_STATUSES: p.status = st
            elif st: warn.append(f"{num}: unknown pamphlet status {st!r} kept as-is"); p.status = st
            p.printed_on = _date(_get(row, "Printed Date")) or p.printed_on
            p.issued_to = circulator(_get(row, "Issued To")) or p.issued_to
            p.issued_on = _date(_get(row, "Issued Date")) or p.issued_on
            p.returned_on = _date(_get(row, "Returned Date")) or p.returned_on
            p.notes = _str(_get(row, "Notes")) or p.notes
            p.print_batch = _str(_get(row, "Print Batch")) or p.print_batch
            p.version_hash = _str(_get(row, "Version Hash")) or p.version_hash
            pam_by_num[num] = p
        db.flush()

    # ---- Signature Sheets --------------------------------------------------------------
    if "Signature Sheets" in wb.sheetnames:
        existing = {(sh.pamphlet.number, sh.sheet_no): sh for sh in db.scalars(select(m.Sheet)).all()}
        for row in _rows(wb["Signature Sheets"], ["Pamphlet #", "Sheet #"]):
            num, k = _str(_get(row, "Pamphlet #")), _int(_get(row, "Sheet #"))
            if not num or not k: continue
            p = pam_by_num.get(num) or db.scalar(select(m.Pamphlet).where(m.Pamphlet.number == num))
            if p is None:
                p = m.Pamphlet(number=num); db.add(p); db.flush(); counts["pamphlets"] += 1; pam_by_num[num] = p
            sh = existing.get((num, k))
            if sh is None:
                sh = m.Sheet(pamphlet=p, sheet_no=k); db.add(sh); counts["sheets"] += 1; existing[(num, k)] = sh
            st = _str(_get(row, "Status"))
            if st: sh.status = st
            sh.circulator = circulator(_get(row, "Circulator")) or sh.circulator
            sh.issued_on = _date(_get(row, "Issued Date")) or sh.issued_on
            sh.returned_on = _date(_get(row, "Returned Date")) or sh.returned_on
            sh.notarized_on = _date(_get(row, "Notarized Date")) or sh.notarized_on
            sh.collected = _int(_get(row, "Collected Signatures")) or 0
            sh.questionable = _int(_get(row, "Questionable")) or 0
            sh.rejected = _int(_get(row, "Rejected")) or 0
            sh.notes = _str(_get(row, "Notes")) or sh.notes
            sh.notary_name = _str(_get(row, "Notary Name")) or sh.notary_name
            sh.notary_commission = _str(_get(row, "Notary Commission #")) or sh.notary_commission
            sh.notary_expiration = _date(_get(row, "Notary Expiration")) or sh.notary_expiration
            dc = _str(_get(row, "Defect Codes"))
            if dc: sh.defect_codes = ",".join(c.strip().upper() for c in re.split(r"[,; ]+", dc) if c.strip())
        db.flush()

    # ---- Issues ---------------------------------------------------------------------
    if "Issues" in wb.sheetnames:
        sheets_by_id = {sh.sheet_id: sh for sh in db.scalars(select(m.Sheet)).all()}
        for row in _rows(wb["Issues"], ["Issue #"]):
            num = _str(_get(row, "Issue #"))
            # The captain's tracker pre-fills I-001..I-300 as empty "Open" rows; a row with no
            # date/pamphlet/sheet/type/notes is a template line, not an issue — skip it.
            if not any(_str(_get(row, c)) for c in ("Date", "Pamphlet #", "Sheet ID", "Issue Type", "Resolution / Notes")):
                continue
            if not num: continue
            i = db.scalar(select(m.Issue).where(m.Issue.number == num))
            if i is None:
                i = m.Issue(number=num); db.add(i); counts["issues"] += 1
            i.opened_on = _date(_get(row, "Date")) or i.opened_on
            pnum = _str(_get(row, "Pamphlet #"))
            if pnum: i.pamphlet = pam_by_num.get(pnum) or db.scalar(select(m.Pamphlet).where(m.Pamphlet.number == pnum))
            sid = _str(_get(row, "Sheet ID"))
            if sid and sid in sheets_by_id: i.sheet = sheets_by_id[sid]
            i.issue_type = _str(_get(row, "Issue Type")) or i.issue_type
            st = _str(_get(row, "Status"))
            if st: i.status = st
            pr = _str(_get(row, "Priority"))
            if pr: i.priority = pr
            i.notes = _str(_get(row, "Resolution / Notes", "Notes")) or i.notes
        db.flush()

    # ---- Filing QA ------------------------------------------------------------------
    if "Filing QA" in wb.sheetnames:
        for n, row in enumerate(_rows(wb["Filing QA"], ["Task"]), 1):
            task = _fix_text(_str(_get(row, "Task")))
            if not task: continue
            t = db.scalar(select(m.QATask).where(m.QATask.task == task))
            if t is None:
                t = m.QATask(task=task); db.add(t); counts["qa_tasks"] += 1
            t.status = _str(_get(row, "Status")) or t.status
            t.owner = _str(_get(row, "Owner")) or t.owner
            t.notes = _fix_text(_str(_get(row, "Notes"))) or t.notes
            t.sort_order = n * 10
        db.flush()

    # ---- EventsShifts → Location + Event -----------------------------------------------
    if "EventsShifts" in wb.sheetnames:
        for row in _rows(wb["EventsShifts"], ["Date", "Location"]):
            lname = _str(_get(row, "Location"))
            if not lname: continue
            slug = _slug(lname)
            loc = db.scalar(select(m.Location).where(m.Location.slug == slug))
            if loc is None:
                loc = m.Location(slug=slug, name=lname, status="planned", public=False); db.add(loc); db.flush(); counts["locations"] += 1
            d, st = _date(_get(row, "Date")), _time(_get(row, "Start"))
            e = db.scalar(select(m.Event).where(m.Event.location_id == loc.id, m.Event.date == d, m.Event.start == st))
            if e is None:
                e = m.Event(location=loc, date=d, start=st, public=False); db.add(e); counts["events"] += 1
            e.end = _time(_get(row, "End")) or e.end
            e.lead = circulator(_get(row, "Event Lead")) or e.lead
            e.volunteers_needed = _int(_get(row, "Volunteers Needed"))
            e.pamphlets_issued = _int(_get(row, "Pamphlets Issued"))
            e.expected_signatures = _int(_get(row, "Expected Signatures"))
            e.notes = _str(_get(row, "Notes")) or e.notes
        db.flush()

    # ---- Dashboard inputs → Settings (label-driven, never fixed cells) -----------------
    if "Dashboard" in wb.sheetnames:
        s = Settings(db)
        labels = {}
        for r in wb["Dashboard"].iter_rows(min_row=1, max_row=40, max_col=2, values_only=True):
            if isinstance(r[0], str) and r[1] not in (None, ""):
                labels[r[0].strip().lower()] = r[1]
        def pick(prefix):
            for k, v in labels.items():
                if k.startswith(prefix.lower()): return v
            return None
        mapping = [("print run", "print_run", _int), ("signature sheets per pamphlet", "sheets_per_pamphlet", _int),
                   ("signature lines per sheet", "rows_per_sheet", _int), ("estimated valid %", "est_valid_rate", float),
                   ("overcollection target %", "overcollect_fraction", float),
                   ("registered voters (written", "registered_voters", _int), ("registered voter count — source", "registered_voters_source", _str),
                   ("registered voter count — date", "registered_voters_date", _date), ("adoption date", "adoption_date", _date),
                   ("election date", "election_date", _date)]
        for prefix, key, conv in mapping:
            v = pick(prefix)
            if v in (None, ""): continue
            if isinstance(v, str) and v.startswith("="): continue      # formula cell with no cached value
            try: val = conv(v)
            except (TypeError, ValueError): continue
            if val in (None, ""): continue
            s.set(key, val); counts["settings"] += 1
        if pick("legal minimum needed") not in (None, ""):
            warn.append("Dashboard 'Legal Minimum Needed' was NOT imported: the registered-voter count is set only from the County Election Board's written figure (admin Settings / config/petition.yaml), never from a spreadsheet number.")
        log.warning(warn[-1]) if warn else None

    db.commit()
    return counts


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path")
    ap.add_argument("--database-url", default=None)
    a = ap.parse_args(argv)
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    from app.db import make_engine, init_db
    from sqlalchemy.orm import sessionmaker
    eng = make_engine(a.database_url); init_db(eng)
    with sessionmaker(bind=eng, expire_on_commit=False)() as db:
        c = import_tracker(a.path, db)
    for k, v in c.items():
        if k != "warnings": print(f"{k:12} {v}")
    for w in c["warnings"]: print("WARNING:", w)


if __name__ == "__main__":
    main()
